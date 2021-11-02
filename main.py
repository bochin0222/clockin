import sys
import requests
import base64
import numpy as np
import json
import time
import datetime

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import cv2
from PyQt5 import uic
from PyQt5.QtWidgets import QGridLayout,QScrollArea,QWidget, QLabel, QApplication, QMainWindow,QVBoxLayout,QPushButton
from PyQt5.QtCore import QThread, Qt, pyqtSignal, pyqtSlot, QRect, QTimer,QTime,QSize,QDate
from PyQt5.QtGui import QFont,QImage, QPixmap, QPainter, QPen, QGuiApplication, QIcon, QPalette, QBrush, QMovie

from media import image
import face_recognize

host = "20.112.4.234:8080"

form_class = uic.loadUiType("ui/login.ui")[0]
regist_class = uic.loadUiType("ui/register.ui")[0]
select_class = uic.loadUiType("ui/select.ui")[0]
manage_class = uic.loadUiType("ui/management.ui")[0]
recognize_class = uic.loadUiType("ui/recognize.ui")[0]
user_regist_class = uic.loadUiType("ui/user_register.ui")[0]
face_regist_class = uic.loadUiType("ui/face_register.ui")[0]
all_records_class = uic.loadUiType("ui/all_records.ui")[0]
salary_report_class = uic.loadUiType("ui/salary_report.ui")[0]
all_users_class = uic.loadUiType("ui/all_users.ui")[0]
user_profile_class = uic.loadUiType("ui/user_profile.ui")[0]

th = face_recognize.Thread()
th.start()

class MainWindow(QMainWindow, form_class,):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.success_widget.hide()
        self.registButton.clicked.connect(self.registButtonClicked)
        self.login_button.clicked.connect(self.loginButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked) 

    def checkform(self,account,password):
        if account is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif password is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False

        else:
            return True        

    def loginButtonClicked(self):

        account = self.account.text()
        password = self.password.text()
        check_result = self.checkform(account,password)                                                                                 

        if check_result:
            service = 'login'

            send_dict = {}
            send_dict['account'] = account
            send_dict['password'] = password
            res = requests.post('http://%s/%s'%(host,service), json=send_dict)
            
            response = json.loads(res.content.decode("utf-8"))
            print(response)

            if response["status"]:
                self.select = SelectWindow(account)
                timer = QTimer()
                timer.singleShot(1,self.close)
                self.select.show()
            else:
                self.alert_text.setText("帳號或密碼錯誤")
                self.success_widget.show()

    def registButtonClicked(self):
        self.register = RegistWindow()
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.register.show()
    
    def comfirmButtonClicked(self):
        self.success_widget.hide()

          

class RegistWindow(QWidget, regist_class):
    def __init__(self):
        super(RegistWindow, self).__init__()
        self.setupUi(self)
        self.alert_text.hide()
        self.success_widget.hide()
        self.login_button.clicked.connect(self.loginButtonClicked)
        self.registButton.clicked.connect(self.registButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked)
    def check_form(self,company_name,account,password,comfirm_password,email):
        if company_name is "":
            self.alert_text.setText("欄位不能為空")
            self.alert_text.show()
            self.company_name.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        if account is "":
            self.alert_text.setText("欄位不能為空")
            self.alert_text.show()
            self.account.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        if password is "":
            self.alert_text.setText("欄位不能為空")
            self.alert_text.show()
            self.password.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        if comfirm_password is "":
            self.alert_text.setText("欄位不能為空")
            self.alert_text.show()
            self.comfirm_password.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        if email is "":
            self.alert_text.setText("欄位不能為空")
            self.alert_text.show()
            self.email.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        if comfirm_password != password:
            self.alert_text.setText("密碼不一樣")
            self.alert_text.show()
            self.comfirm_password.setStyleSheet("border: 2px solid red;\n"
                                        "background-color: rgb(255, 255, 255);")
            return False

        return True  

    def loginButtonClicked(self):
        self.login = MainWindow()
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.login.show()
    
    def comfirmButtonClicked(self):
        self.login = MainWindow()
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.login.show()
    
    def registButtonClicked(self):
        self.alert_text.hide()
        self.company_name.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.account.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.password.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.comfirm_password.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.email.setStyleSheet("background-color: rgb(255, 255, 255);")
        company_name = self.company_name.text()
        account = self.account.text()
        password = self.password.text()
        comfirm_password = self.comfirm_password.text()
        email = self.email.text()
        
        check_result = self.check_form(company_name, account, password, comfirm_password, email)

        if check_result:
            service = 'company_register'
            send_dict = {}
            send_dict['account'] = account
            send_dict['password'] = password
            send_dict['mail'] = email
            send_dict['workspace'] = company_name
            send_dict['third_party'] = 'None'

            res = requests.post('http://%s/%s'%(host,service), json=send_dict)
            print(res.content.decode("utf-8"))
            response = json.loads(res.content.decode("utf-8"))

            if response["status"]:
                self.success_widget.show()


          

class SelectWindow(QWidget, select_class):
    def __init__(self,account):
        super(SelectWindow, self).__init__()
        self.setupUi(self)
        self.account = account
        self.logoutButton.clicked.connect(self.logoutButtonClicked)
        self.manageButton.clicked.connect(self.manageButtonClicked)
        self.clockinButton.clicked.connect(self.clockinButtonClicked)
        self.displayTime()
        timer = QTimer(self)
        timer.timeout.connect(self.displayTime)
        timer.start(1000)
    
    def displayTime(self):
        currntTime = QTime.currentTime()
        displayText = currntTime.toString('hh:mm')
        self.clock.setText(displayText)

    def logoutButtonClicked(self):
        service = 'logout'

        send_dict = {}
        send_dict['account'] = self.account
        res = requests.post('http://%s/%s'%(host,service), json=send_dict)
        print(res.content.decode("utf-8") )
        self.login = MainWindow()
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.login.show()

    def manageButtonClicked(self):
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.manage.show() 

    def clockinButtonClicked(self):
        self.recognize = RecognizeWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.recognize.show()      

class ManageWindow(QWidget, manage_class):
    def __init__(self,account):
        super(ManageWindow, self).__init__()
        self.setupUi(self)
        self.account = account 
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.addUserButton.clicked.connect(self.adduserButtonClicked)
        self.recordButton.clicked.connect(self.recordButtonClicked)
        self.reportButton.clicked.connect(self.reportButtonClicked)
        self.accountButton.clicked.connect(self.accountButtonClicked)

    def adduserButtonClicked(self):
        self.userRegist = UserRegistWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.userRegist.show()
    
    def reportButtonClicked(self):
        self.report = SalaryReportWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.report.show()
    
    def recordButtonClicked(self):
        service = 'user_record'

        send_dict = {}
        send_dict['account'] = self.account

        res = requests.post('http://%s/%s' % (host, service), json=send_dict)
        response = json.loads(res.content.decode("utf-8"))
        if response["status"]:
            self.all_records = AllRecordsWindow(self.account,response["result"])
            timer = QTimer()
            timer.singleShot(1,self.close)
            self.all_records.show()
    
    def accountButtonClicked(self):
        service = 'get_user_profile'

        send_dict = {}
        send_dict['account'] = self.account
        res = requests.post('http://%s/%s' % (host, service), json = send_dict)
        response = json.loads(res.content.decode("utf-8"))
        if response["status"]:
            self.all_users = AllUsersWindow(self.account,response["result"])
            timer = QTimer()
            timer.singleShot(1,self.close)
            self.all_users.show()
            
        


    def returnButtonClicked(self):
        self.select = SelectWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.select.show()

class UserRegistWindow(QWidget, user_regist_class):
    def __init__(self,account):
        super(UserRegistWindow, self).__init__()
        self.setupUi(self)
        self.success_widget.hide()
        self.account = account
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.faceRegistButton.clicked.connect(self.faceRegistButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked)

    def checkform(self,user_name,phone,email,birthday,maleButton,femaleButton,employeeButton,managerButton):
        if user_name is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif phone is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif email is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif  maleButton.isChecked() is False and  femaleButton.isChecked() is False:
            self.alert_text.setText("性別沒選擇！")
            self.success_widget.show()
            return False
        
        elif employeeButton.isChecked() is False and  managerButton.isChecked() is False:
            self.alert_text.setText("身份沒選擇！")
            self.success_widget.show()
            return False


        else:
            return True  

    def returnButtonClicked(self):
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.manage.show()
    
    def faceRegistButtonClicked(self):

        user_name = self.name.text()
        phone = self.phone.text()
        email = self.email.text()
        birthday = self.birthday.date().toString(Qt.ISODate)
        
        check_result = self.checkform(user_name,phone,email,birthday,self.maleButton,self.femaleButton,self.employeeButton,self.managerButton)

        if check_result:
            sex = "male" if self.maleButton.isChecked() is True else "female"
            manager = False if self.employeeButton.isChecked() is True else False
            
            send_dict = {}
            send_dict['account'] = self.account
            send_dict['name'] = user_name
            send_dict['phone'] = phone
            send_dict['mail'] = email
            send_dict['sex'] = sex
            send_dict['birthday'] = birthday
            send_dict['manager'] = manager
            
            th.mode = "register"
            self.face = FaceRegistWindow(send_dict)
            timer = QTimer()
            timer.singleShot(1,self.close)
            self.face.show()
    
    def comfirmButtonClicked(self):
        self.success_widget.hide()

class FaceRegistWindow(QWidget, face_regist_class):
    def __init__(self,send_dict):
        super(FaceRegistWindow, self).__init__()
        self.setupUi(self)
        self.success_widget.hide()
        self.send_dict = send_dict
        self.account = self.send_dict['account']
        th.start_recognize = True
        th.regist_frame.connect(self.setImage)
        th.count_text.connect(self.setStatus)
        th.regist_img.connect(self.get_cropped_image)
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked)

    @pyqtSlot(QImage)
    def setImage(self, image):
        self.frame.setPixmap(QPixmap.fromImage(image))

    @pyqtSlot(np.ndarray)
    def get_cropped_image(self, image):
        _,encode_image = cv2.imencode('.jpg',image)
        self.send_dict['cropimage'] = str(base64.b64encode(encode_image))
        self.send_dict['landmarks'] = th.landmarks.tolist()
        self.upload_user_file(self.send_dict)

    @pyqtSlot(str)
    def setStatus(self, text):
        self.status_text.setText(text)

    def upload_user_file(self,send_dict):
        service = 'user_reigster'

        res = requests.post('http://%s/%s'%(host,service), json=send_dict)
        print(res.content.decode("utf-8"))
        response = json.loads(res.content.decode("utf-8"))
        print(response)
        if response["status"]:
            self.alert_text.setText("註冊成功")
            self.success_widget.show()

    def returnButtonClicked(self):
        th.mode = "identify"
        th.start_recognize = False
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.deleteLater)
        self.manage.show()

    def comfirmButtonClicked(self):
        th.mode = "identify"
        th._stop()
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.deleteLater)
        self.manage.show()

class RecognizeWindow(QWidget, recognize_class):
    def __init__(self,account):
        super(RecognizeWindow, self).__init__()
        self.setupUi(self)
        self.account = account
        th.camera_frame.connect(self.setImage)
        th.status_text.connect(self.setStatus)
        th.identify_img.connect(self.get_cropped_image)
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.notMeButton.clicked.connect(self.notMeButtonClicked)
        self.workButton.clicked.connect(self.workButtonClicked)
        self.offworkButton.clicked.connect(self.offworkButtonClicked)
        self.recordButton.clicked.connect(self.recordButtonClicked)
        self.yesButton.clicked.connect(self.yesButtonClicked)
        self.noButton.clicked.connect(self.noButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked)
        th.start_recognize = True
        self.received_dict = {}
        self.widget.hide()
        self.widget_2.hide()
        self.widget_3.hide()
        self.scrollArea.hide()
        self.status = None
        self.name = ""
        self.displayTime()
        timer=QTimer(self)
        timer.timeout.connect(self.displayTime)
        timer.start(1000)
    
    def displayTime(self):
        currntTime = QTime.currentTime()
        displayText = currntTime.toString('hh:mm')
        self.clock.setText(displayText)
    
    @pyqtSlot(QImage)
    def setImage(self, image):
        self.frame.setPixmap(QPixmap.fromImage(image))

    @pyqtSlot(str)
    def setStatus(self, text):
        self.status_text.setText(text)
    
    @pyqtSlot(np.ndarray)
    def get_cropped_image(self, image):
        
        _,encode_image = cv2.imencode('.jpg',image)
        self.start_identify(encode_image)
    

    def start_identify(self,image):
        send_dict ={}
        service = 'identify'
        send_dict['cropimage'] = str(base64.b64encode(image.copy()))
        send_dict['landmarks'] = th.landmarks.tolist()
        send_dict['account'] = self.account
        res = requests.post('http://%s/%s'%(host,service), json=send_dict)
        response = json.loads(res.content.decode("utf-8"))
        
        self.received_dict = response["result"]
        if response["status"]:
            self.show_result(self.received_dict)
        else:
            th.start_recognize = True

    def process_image(self,image):
        image = cv2.resize(image,(150,150))
        
        rgbImage = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        h, w, ch = rgbImage.shape
        bytesPerLine = ch * w
        convertToQtFormat = QImage(rgbImage.data, w, h, bytesPerLine, QImage.Format_RGB888)
        
        p = convertToQtFormat.scaled(150, 150)
        return p

    def decode_image(self,base64_buffer):

        raw_image = base64_buffer[2:].encode()
        raw_image = base64.decodebytes(raw_image)
        raw_image = np.frombuffer(raw_image,np.uint8)
        image = cv2.imdecode(raw_image,cv2.IMREAD_COLOR)
        return image

    def show_result(self,result_dict):
        th._stop()
        self.name = result_dict["username"]
        image = result_dict["image"]
        image = self.decode_image(image)
        image = self.process_image(image) 
        self.user_image.setPixmap(QPixmap.fromImage(image))
        print(result_dict["username"],type(result_dict["username"]))
        self.user_name.setText(self.name)
        self.widget.show()
    
    def set_record_widgets(self,index,dict):
         
        widget = QWidget(self.scrollAreaWidgetContents)
        widget.setGeometry(QRect(10, 10 + 80*index, 231, 71))
        
        date_ = datetime.datetime.strptime(dict["date"].split(" ")[0],'%Y/%m/%d')
        week = {0:"(一)",1:"(二)",2:"(三)",3:"(四)",4:"(五)",5:"(六)",6:"(日))"}
        weekday = week[date_.weekday()]
        date = QLabel(widget)
        date.setGeometry(QRect(33, 7, 122, 26))
        font = QFont()
        font.setFamily("Roboto")
        date.setFont(font)
        date.setAlignment(Qt.AlignCenter)
        date.setText(dict["date"].split(" ")[0] + " " + weekday)

        time_text = QLabel(widget)
        time_text.setGeometry(QRect(50, 37, 76, 26))
        font.setFamily("Roboto")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        time_text.setFont(font)
        time_text.setAlignment(Qt.AlignCenter)
        time_text.setText(dict["date"].split(" ")[1].split(":")[0] + ":" +dict["date"].split(" ")[1].split(":")[1])

        clockin_text = QLabel(widget)
        clockin_text.setGeometry(QRect(160, 24, 55, 25))
        font = QFont()
        font.setFamily("Roboto")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        clockin_text.setFont(font)
        clockin_text.setAlignment(Qt.AlignCenter)
        if dict["status"] == "ON":
            clockin_text.setText("上班")
            widget.setStyleSheet("background-color: rgb(174, 255, 160);\n"
                                    "border-radius: 29px;\n")
        else:
            clockin_text.setText("下班")
            widget.setStyleSheet("background-color: rgb(255, 205, 199);\n"
                                    "border-radius: 29px;\n")

        widget.show()

    def returnButtonClicked(self):
        th._stop()
        self.select = SelectWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.deleteLater)
        self.select.show()
        
    def workButtonClicked(self):
        self.status = "ON"
        self.comfirm_text.setText("確定打上班卡？")
        self.widget_2.show()

    def offworkButtonClicked(self):
        self.status = "OFF"
        self.comfirm_text.setText("確定打下班卡？")
        self.widget_2.show()
    
    def yesButtonClicked(self):
        service = 'clockin'

        send_dict = {}
        send_dict['account'] = self.account
        send_dict['userid'] = self.received_dict["user_id"]
        send_dict['recordID'] = self.received_dict["record_id"]
        send_dict['status'] = self.status
        send_dict['date'] = datetime.datetime.today().strftime("%Y/%m/%d %H:%M:%S") 
        res = requests.post('http://%s/%s'%(host,service), json=send_dict)
        response = json.loads(res.content.decode("utf-8"))
        if response["status"]:
            self.alert_text.setText("您於{} 打卡成功!!".format(send_dict['date']))
            self.widget_3.show()
    
    def comfirmButtonClicked(self):
        th._stop()
        self.select = SelectWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.deleteLater)
        self.select.show()

    def noButtonClicked(self):
        self.widget_2.hide()

    def recordButtonClicked(self):

        service = 'user_record'

        send_dict = {}
        send_dict['account'] = self.account

        res = requests.post('http://%s/%s' % (host, service), json=send_dict)

        response = json.loads(res.content.decode("utf-8"))
        print(response["status"])
        if response["status"]:
            user_records = []
            response["result"].reverse()
            for record in response["result"]:
                if record['user'] == self.name:
                    user_records.append(record)
            minimum_Size = 81 * len(user_records) if len(user_records) > 5 else 0
            self.scrollAreaWidgetContents.setMinimumSize(0, minimum_Size)

            for i in self.scrollAreaWidgetContents.findChildren(QWidget):
                i.setParent(None)
        
            for i,record in enumerate(user_records):
                self.set_record_widgets(i, record)
            self.scrollArea.show()

    def notMeButtonClicked(self):

        th.start_recognize = True
        self.widget.hide()

class AllRecordsWindow(QWidget, all_records_class):
    def __init__(self,account,records):
        super(AllRecordsWindow, self).__init__()
        self.setupUi(self)
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.pushButton.clicked.connect(self.pushButtonClicked)
        self.records = records
        self.records.reverse()
        self.account = account
        self.minimum_Size = 91 * len(self.records) if len(self.records) > 5 else 0
        self.scrollAreaWidgetContents_2.setMinimumSize(0, self.minimum_Size)
        for i,record in enumerate(self.records):
            self.set_record_widgets(i, record,"initial")


    
    def process_image(self,image):
        # image = cv2.resize(image,(48,58))
        
        rgbImage = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        h, w, ch = rgbImage.shape
        bytesPerLine = ch * w
        convertToQtFormat = QImage(rgbImage.data, w, h, bytesPerLine, QImage.Format_RGB888)
        
        p = convertToQtFormat.scaled(48, 58)
        return p
    
    def decode_image(self,base64_buffer):
        raw_image = base64_buffer[2:].encode()
        raw_image = base64.decodebytes(raw_image)
        raw_image = np.frombuffer(raw_image,np.uint8)
        image = cv2.imdecode(raw_image,cv2.IMREAD_COLOR)
        return image

    def set_record_widgets(self,index,dict,status):
         
        widget = QWidget(self.scrollAreaWidgetContents_2)
        widget.setGeometry(QRect(13, 12 + 87*index, 301, 81))
        
        user_img = self.decode_image(dict["image"])
        user_img = self.process_image(user_img)

        img = QLabel(widget)
        img.setGeometry(QRect(25, 12, 48, 58))
        img.setPixmap(QPixmap.fromImage(user_img))
        
        date_ = datetime.datetime.strptime(dict["date"].split(" ")[0],'%Y/%m/%d')
        week = {0:"(一)",1:"(二)",2:"(三)",3:"(四)",4:"(五)",5:"(六)",6:"(日))"}
        weekday = week[date_.weekday()]
        date = QLabel(widget)
        date.setGeometry(QRect(82, 11, 122, 26))
        font = QFont()
        font.setFamily("Roboto")
        date.setFont(font)
        date.setAlignment(Qt.AlignCenter)
        date.setText(dict["date"].split(" ")[0] + " " + weekday)

        time_text = QLabel(widget)
        time_text.setGeometry(QRect(99, 40, 76, 26))
        font.setFamily("Roboto")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        time_text.setFont(font)
        time_text.setAlignment(Qt.AlignCenter)
        time_text.setText(dict["date"].split(" ")[1].split(":")[0] + ":" +dict["date"].split(" ")[1].split(":")[1])

        clockin_text = QLabel(widget)
        clockin_text.setGeometry(QRect(200, 10, 81, 61))
        font = QFont()
        font.setFamily("Roboto")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        clockin_text.setFont(font)
        clockin_text.setAlignment(Qt.AlignCenter)
        if dict["status"] == "ON":
            clockin_text.setText("上班")
            widget.setStyleSheet("background-color: rgb(174, 255, 160);\n"
                                    "border-radius: 29px;\n")
        else:
            clockin_text.setText("下班")
            widget.setStyleSheet("background-color: rgb(255, 205, 199);\n"
                                    "border-radius: 29px;\n")
        if status == "not initial":
            widget.show()

    
    def pushButtonClicked(self):
        self.records.reverse()
        # self.scrollAreaWidgetContents_2.deleteLater()
        for i in self.scrollAreaWidgetContents_2.findChildren(QWidget):
            i.setParent(None)
       
        for i,record in enumerate(self.records):
            self.set_record_widgets(i, record,"not initial")
        
        
    def returnButtonClicked(self):
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.manage.show()

class SalaryReportWindow(QWidget, salary_report_class):
    def __init__(self,account):
        super(SalaryReportWindow, self).__init__()
        self.account = account
        self.setupUi(self)
        self.widget_2.hide()
        self.start_date_str = ""
        self.end_date_str = ""
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.outputButton.clicked.connect(self.outputButtonClicked)
        self.cancelButton.clicked.connect(self.cancelButtonClicked)
        self.yesButton.clicked.connect(self.yesButtonClicked)
    
    def output_excel(self,datas,start_date,end_date):
        wb = Workbook()
        ws = wb.active

        title = ['姓名','工時']
        ws.append(title)

        for person in datas.keys():
            ws.append([person,datas[person]])
        
        for col in range(1,5):
            char = get_column_letter(col)
            ws[char + '1'].font = Font(bold=True)
        
        wb.save("salary_report/salary_report_{}_{}.xlsx".format(start_date,end_date))
        print("save excel successfully!")


    def outputButtonClicked(self):
        self.start_date_str = self.start_date.date().toString("yyyy/MM/dd")
        self.end_date_str = self.end_date.date().toString("yyyy/MM/dd")
        cal_start_date = datetime.datetime.strptime(self.start_date_str,'%Y/%m/%d')
        cal_end_date = datetime.datetime.strptime(self.end_date_str,'%Y/%m/%d')
        delta = cal_end_date - cal_start_date
        if delta.days >= 0:
            self.date_range_text.setText("日期區間 : {} ~ {}".format(self.start_date_str,self.end_date_str))
            self.widget_2.show()

    def yesButtonClicked(self):
        service = 'cal_working_hours'

        send_dict = {}
        send_dict['account'] = self.account
        send_dict['starttime'] = self.start_date_str
        send_dict['endtime'] = self.end_date_str
        res = requests.post('http://%s/%s' % (host, service), json = send_dict)
        response = json.loads(res.content.decode("utf-8"))
        if response["status"]:
            start_date,end_date = "",""
            for i in self.start_date_str.split("/"):
                start_date = start_date + i
            for i in self.end_date_str.split("/"):
                end_date = end_date + i
            self.output_excel(response["result"],start_date,end_date)
            self.widget_2.close()


    def cancelButtonClicked(self):
        self.widget_2.close()

    def returnButtonClicked(self):
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.manage.show()

class AllUsersWindow(QWidget, all_users_class):
    def __init__(self,account,users):
        super(AllUsersWindow, self).__init__()
        self.account = account
        self.setupUi(self)
        self.users = users
        self.minimum_Size = 45 * len(self.users) if len(self.users) > 9 else 0
        self.scrollAreaWidgetContents_2.setMinimumSize(0, self.minimum_Size)
        for i,user in enumerate(self.users):
            self.set_all_users(i, user)
        self.returnButton.clicked.connect(self.returnButtonClicked)
        
    def set_all_users(self,index,user):
        widget = QWidget(self.scrollAreaWidgetContents_2)
        widget.setGeometry(QRect(0, 6 + 45*index, 320, 39))
        
        name = QLabel(widget)
        name.setGeometry(QRect(7, 6, 91, 26))
        font = QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setFamily("Roboto")
        name.setFont(font)
        name.setAlignment(Qt.AlignCenter)
        name.setText(user["name"])

        email = QLabel(widget)
        email.setGeometry(QRect(102, 14, 133, 11))
        font = QFont()
        font.setPointSize(6)
        font.setFamily("Roboto")
        email.setFont(font)
        email.setAlignment(Qt.AlignCenter)
        email.setText(user["mail"])

        pushButton = QPushButton(widget)
        pushButton.setGeometry(QRect(272, 7, 24, 24))
        icon = QIcon()
        icon.addPixmap(QPixmap(":/newPrefix/bx_bx-edit.png"), QIcon.Normal,QIcon.Off)
        pushButton.setIcon(icon)
        pushButton.setIconSize(QSize(24, 24))
        pushButton.setFlat(True)
        pushButton.clicked.connect(lambda:self.pushButtonClicked(user))

    def pushButtonClicked(self,user_file):
        self.profile = UserProfileWindow(self.account,self.users,user_file)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.profile.show()
    
    def returnButtonClicked(self):
        self.manage = ManageWindow(self.account)
        timer = QTimer()
        timer.singleShot(1,self.close)
        self.manage.show()


class UserProfileWindow(QWidget, user_profile_class):
    def __init__(self,account,users,user_file):
        print(user_file)
        super(UserProfileWindow, self).__init__()
        self.setupUi(self)
        self.account = account
        self.users = users
        self.user_file = user_file
        self.success_widget.hide()
        self.set_user_profile()
        self.returnButton.clicked.connect(self.returnButtonClicked)
        self.saveButton.clicked.connect(self.saveButtonClicked)
        self.comfirmButton.clicked.connect(self.comfirmButtonClicked)
    
    def set_user_profile(self):
        self.name.setText(self.user_file["name"])
        self.phone.setText(self.user_file["phone"])
        self.email.setText(self.user_file["mail"])
        birthday = self.user_file["birthday"].split("-")
        self.birthday.setDate(QDate(int(birthday[0]), int(birthday[1]), int(birthday[2])))
        if self.user_file["sex"] == 'male':
            self.maleButton.setChecked(True)
        elif self.user_file["sex"] == 'female':
            self.femaleButton.setChecked(True)
        if self.user_file["manager"]:
            self.managerButton.setChecked(True)
        else:
            self.employeeButton.setChecked(True)
    
    def checkform(self,user_name,phone,email,birthday,maleButton,femaleButton,employeeButton,managerButton):
        if user_name is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif phone is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif email is "":
            self.alert_text.setText("欄位不能為空")
            self.success_widget.show()
            return False
        
        elif  maleButton.isChecked() is False and  femaleButton.isChecked() is False:
            self.alert_text.setText("性別沒選擇！")
            self.success_widget.show()
            return False
        
        elif employeeButton.isChecked() is False and  managerButton.isChecked() is False:
            self.alert_text.setText("身份沒選擇！")
            self.success_widget.show()
            return False


        else:
            return True


    def saveButtonClicked(self):
        user_name = self.name.text()
        phone = self.phone.text()
        email = self.email.text()
        birthday = self.birthday.date().toString(Qt.ISODate)
        
        check_result = self.checkform(user_name,phone,email,birthday,self.maleButton,self.femaleButton,self.employeeButton,self.managerButton)
        if check_result:
            sex = "male" if self.maleButton.isChecked() is True else "female"
            manager = False if self.employeeButton.isChecked() is True else False
            
            send_dict = {}
            send_dict['account'] = self.account
            send_dict['name'] = user_name
            send_dict['phone'] = phone
            send_dict['mail'] = email
            send_dict['sex'] = sex
            send_dict['birthday'] = birthday
            send_dict['manager'] = manager

            service = 'edit_user_profile'

            res = requests.post('http://%s/%s'%(host,service), json=send_dict)
            response = json.loads(res.content.decode("utf-8"))
            if response["status"]:
                self.alert_text.setText("儲存成功！")
                self.success_widget.show()
            
    def returnButtonClicked(self):
        service = 'get_user_profile'

        send_dict = {}
        send_dict['account'] = self.account
        res = requests.post('http://%s/%s' % (host, service), json = send_dict)
        response = json.loads(res.content.decode("utf-8"))
        if response["status"]:
            self.all_users = AllUsersWindow(self.account,response["result"])
            timer = QTimer()
            timer.singleShot(1,self.close)
            self.all_users.show()

    def comfirmButtonClicked(self): 
        self.success_widget.hide()


        
       

        

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())